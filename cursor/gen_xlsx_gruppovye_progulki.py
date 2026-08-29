#!/usr/bin/env python3
"""ЕПК групповые прогулки — xlsx в формате выгрузки Директа (704503370)."""

from __future__ import annotations

from copy import copy
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

TEMPLATE = Path(
    "/home/ubuntu/.cursor/projects/workspace/uploads/704503370_b81e.xlsx"
)
OUT = Path(__file__).resolve().parent / "EPK_Gruppovye_progulki_yahta_Direkt_2026-08-11.xlsx"

LANDING = "https://boat-sochi.ru/progulki_na_yacht"
LANDING_UTM = (
    LANDING
    + "?utm_source=yandex&utm_medium=cpc"
    + "&utm_campaign=epk_gruppovye_yahta_progulki"
)
PHONE = "+7 918 304-40-00"
DISPLAY = "Прогулка-на-яхте"

# --- 8 быстрых ссылок (всегда 8) ---
# Описания: 2 строки внутри одной ссылки = \n (НЕ ||).
# || — только разделитель между 8 ссылками в колонках BC/BD/BE.
SITELINK_TITLES = [
    "Расписание слотов",
    "Цена от 1 800 ₽",
    "Закат в 18:00",
    "К дельфинам",
    "Причал Сириус",
    "Группа до 11",
    "Забронировать",
    "Как добраться",
]
SITELINK_DESCS = [
    "С 9:00 до 18:00\nКаждый день",
    "Дневные слоты\nЗакат 2 500 ₽",
    "Билет 2 500 ₽\nПрогулка 1,5 часа",
    "Часто встречаются\nЖивая природа",
    "Имеретинский порт\nЛиния 2",
    "Парусная яхта\n1,5 часа в море",
    "Звонок или MAX\nОтветим быстро",
    "Парусная, 1\nСириус, линия 2",
]
SITELINK_URLS = [
    f"{LANDING}?utm_content=sl_schedule",
    f"{LANDING}?utm_content=sl_price",
    f"{LANDING}?utm_content=sl_sunset",
    f"{LANDING}?utm_content=sl_dolphins",
    f"{LANDING}?utm_content=sl_pier",
    f"{LANDING}?utm_content=sl_group",
    f"{LANDING}?utm_content=sl_book",
    f"{LANDING}?utm_content=sl_map",
]

CLARIFICATIONS = "||".join(
    [
        "1,5 часа на море",
        "группа до 11 человек",
        "от 1 800 рублей",
        "закат 2 500 ₽",
        "Имеретинский порт",
        "парусная яхта",
        "ребёнок до 5 бесплатно",
        "слоты каждый день",
    ]
)

CAMPAIGN_MINUS = (
    "-!в архипо осиповке -!в новороссийске -!в туапсе -!во владивостоке "
    "-Воронеж -Кострома -Новосибирск -Петербург -Самара -Чебоксары "
    "-анапа -астрахань -балаклава -барнаул -бархатные сезоны -башкортостан "
    "-без капитана -владивосток -геленджик -дивноморск -евпатория -завидово "
    "-кабардинка -казань -калининград -кисегач -конные -конь -крым -купить "
    "-курортный сбор -лошадь -москва -недорого -нижний -нижний новгород "
    "-новороссийск -океанариум -песок -петрозаводск -питер -пляж -поляна "
    "-пошехонье -приморскогокрая -продать -псков -ремонт -ростов -рыбинск "
    "-санкт петербург -севастополь -спб -стамбул -таганрог -темрюк -тольятти "
    "-трансфер -турция -феодосия -ялта -ярославль "
    "-аренда катера -снять яхту -снять катер -вип -vip -премиум -алексум "
    "-tigger -тиггер -сутки -на сутки -круиз -рыбалка -трофейн -гидроцикл "
    "-теплоход -7500 -частная аренда -дельфинарий -шоу дельфинов "
    "-плавать с дельфинами -вакансия -работа -купить яхту -продажа яхты "
    "-чертеж -симулятор -игра -бесплатно"
)

GROUP_MINUS = (
    "-аренда катера -аренда яхты целиком -снять катер -снять яхту "
    "-каютный катер -на сутки -рыбалка -теплоход -дельфинарий -океанариум "
    "-вип -vip -алексум -tigger -гидроцикл -7500 -вакансия -купить -продать"
)

# Мусорные площадки РСЯ (из аудитов + клики без конверсий в выгрузках)
JUNK_SITES = [
    # игры / мессенджеры
    "com.imo.android.imoim",
    "com.imo.android.imoimhd",
    "com.oakever.tiletrip",
    "com.kayac.ball_run",
    "com.outfit7.mytalkingtomfriends",
    "com.outfit7.talkingtomgoldrun",
    "com.outfit7.herodash",
    "com.outfit7.mytalkingangela2",
    "com.europosit.pixelcoloring",
    "coloring.color.number.happy.paint.art.drawing.puzzle",
    "com.fingerlab.word.blockpuzzles",
    "com.game.spongeart",
    "com.goods.sorting.games.triple.match3d.puzzle",
    "com.kadka.forknsausage",
    "com.openmygame.games.android.jigsawpuzzle",
    "com.orangeapps.piratetreasures.2",
    "com.solitaire.klondike.patience.ocean.aquarium",
    "com.starplay.spider.fighter.openworld",
    "com.zm.watersort",
    "com.screw3d.match.nuts.bolts.pin.jam.away.puzzle",
    "com.pd.vehiclemasters",
    "com.abi.cook.chill",
    "com.grif.vmp",
    "funvent.tilepark",
    "game.water.sort.puzzle.android",
    "games.vaveda.militaryoverturn",
    "io.supercent.linkedcubic",
    "io.voodoo.paper2",
    "kidultlovin.word.zen",
    "puzzle.blockpuzzle.cube.relax",
    "teskin.jewels.planet",
    "traffic.parking.jam.escape.car3d.games",
    "ai.character.app",
    # инфо / погода / агрегаторы с 0 CR
    "m.pogoda.yandex.ru",
    "dzen.ru",
    "otzovik.com",
    "nashaspravka.ru",
    "zoon.ru",
    "orgs.biz",
    "sochi1.ru",
    "m.video.yandex.ru",
    "m.images.yandex.ru",
    # соцсети/классифайды как РСЯ-площадки (низкий интент)
    "com.vkontakte.android",
    "com.vk.vkclient",
    "com.avito.android",
    "avito.ru",
    "ru.ok.android",
]

HEADLINES = [
    "Групповая прогулка на яхте от 1 800 ₽",
    "Парусная яхта Сириус - 1,5 часа",
    "Увидеть дельфинов - Имеретинский порт",
    "Морская прогулка к дельфинам, Сириус",
    "Закат с яхты - слот 18:00, 2 500 ₽",
    "Билет на яхту: группа до 11 человек",
    "Яхта из Имеретинского порта, линия 2",
]

TEXTS = [
    "Групповая прогулка 1,5 часа. От 1800 ₽, закат 2500 ₽. Дельфины, Олимпийский парк.",
    "Не снимайте яхту целиком - купите место. До 11 человек. Бронь: " + PHONE,
    "Сириус, линия 2. Слоты 9:00-18:00. Ребенок до 5 лет на яхту - бесплатно.",
]

GROUPS = [
    (
        "G1 Групповая прогулка",
        1,
        [
            '"групповая прогулка на яхте сочи"',
            '"групповая прогулка на яхте сириус"',
            '"прогулка на яхте сочи цена"',
            '"морская прогулка на яхте сочи"',
            '"морская прогулка на парусной яхте"',
            '"билет на яхту сочи"',
            '"билет на яхту сириус"',
            '"прогулка на яхте за человека"',
            '"яхта сочи 1800"',
            '"морская прогулка 1.5 часа сочи"',
        ],
    ),
    (
        "G2 Дельфины Сириус",
        2,
        [
            '"прогулка к дельфинам сочи"',
            '"прогулка к дельфинам сириус"',
            '"увидеть дельфинов сочи"',
            '"дельфины в море сочи экскурсия"',
            '"яхта к дельфинам сочи"',
            '"морская прогулка дельфины адлер"',
            '"дельфины имеретинский порт"',
        ],
    ),
    (
        "G3 Закат на яхте",
        3,
        [
            '"прогулка на яхте на закате сочи"',
            '"закат на яхте сириус"',
            '"яхта закат олимпийский парк"',
            '"вечерняя прогулка на яхте сочи"',
            '"закат с яхты сочи"',
        ],
    ),
    (
        "G4 Сириус / Имеретинский порт",
        4,
        [
            '"морские прогулки сириус"',
            '"морская прогулка имеретинский порт"',
            '"яхта имеретинский порт"',
            '"прогулка на яхте сириус"',
            '"морская прогулка сириус"',
            '"морские прогулки в сириусе"',
        ],
    ),
    (
        "G5 Адлер / Олимпийский парк",
        5,
        [
            '"прогулка на яхте адлер"',
            '"морские прогулки адлер"',
            '"яхта олимпийский парк"',
            '"экскурсия на яхте олимпийский парк"',
            '"морская прогулка олимпийский парк"',
        ],
    ),
]


def len_formula(cell_ref: str) -> str:
    # Excel: """" внутри формулы = символ кавычки для SUBSTITUTE
    return (
        '=IF({c}="","",LEN(SUBSTITUTE(SUBSTITUTE(SUBSTITUTE(SUBSTITUTE('
        'SUBSTITUTE(SUBSTITUTE({c},"!",""),",",""),".",""),";",""),":",""),"""","")))'
    ).format(c=cell_ref)


def check_limits():
    for h in HEADLINES:
        n = len(h.replace("!", "").replace(",", "").replace(".", "").replace(";", "").replace(":", "").replace('"', ""))
        assert n <= 56, (h, n)
    for t in TEXTS:
        n = len(t.replace("!", "").replace(",", "").replace(".", "").replace(";", "").replace(":", "").replace('"', ""))
        assert n <= 81, (t, n)
    for t in SITELINK_TITLES:
        assert len(t) <= 30, t
    assert len(SITELINK_TITLES) == 8
    assert len(SITELINK_DESCS) == 8
    assert len(SITELINK_URLS) == 8
    joined = "||".join(SITELINK_DESCS)
    assert joined.count("||") == 7, joined
    for d in SITELINK_DESCS:
        assert "||" not in d, d
        for line in d.split("\n"):
            assert len(line) <= 30, (line, len(line))


def copy_cell_style(src, dst):
    if src.has_style:
        dst.font = copy(src.font)
        dst.border = copy(src.border)
        dst.fill = copy(src.fill)
        dst.number_format = src.number_format
        dst.protection = copy(src.protection)
        dst.alignment = copy(src.alignment)


def clear_data_rows(ws, start_row=12):
    if ws.max_row >= start_row:
        ws.delete_rows(start_row, ws.max_row - start_row + 1)


def build():
    check_limits()
    wb = openpyxl.load_workbook(TEMPLATE)
    ws = wb["Тексты"]

    # meta
    ws["A6"] = "Предложение текстовых блоков для кампании"
    ws["E7"] = "Единая перфоманс-кампания"
    ws["E8"] = "EPK_Gruppovye_Yahta_Progulki"
    ws["H8"] = "RUB"
    ws["E9"] = CAMPAIGN_MINUS

    # keep header rows 10-11, clear old data
    clear_data_rows(ws, 12)

    # style sample from template was lost after delete — write plain values + formulas
    sitelink_titles = "||".join(SITELINK_TITLES)
    sitelink_descs = "||".join(SITELINK_DESCS)
    sitelink_urls = "||".join(SITELINK_URLS)

    row = 12
    for group_name, group_num, phrases in GROUPS:
        for idx, phrase in enumerate(phrases):
            first_in_group = idx == 0
            # A–H
            ws.cell(row, 1, "-")
            ws.cell(row, 2, "Комбинаторное")
            ws.cell(row, 3, None)  # ID группы — новая РК
            ws.cell(row, 4, group_name)
            ws.cell(row, 5, group_num)
            ws.cell(row, 6, None)  # ID фразы
            ws.cell(row, 7, phrase)
            ws.cell(row, 8, None)  # ID объявления
            # I–K empty for combinator
            for col in (9, 10, 11):
                ws.cell(row, col, None)
            # L M N length formulas for I J K
            ws.cell(row, 12, len_formula(f"I{row}"))
            ws.cell(row, 13, len_formula(f"J{row}"))
            ws.cell(row, 14, len_formula(f"K{row}"))
            # O–U headlines (7)
            for i, h in enumerate(HEADLINES):
                ws.cell(row, 15 + i, h)
            # V–X texts (3)
            for i, t in enumerate(TEXTS):
                ws.cell(row, 22 + i, t)
            # Y–AE length headlines, AF–AH length texts
            for i in range(7):
                ws.cell(row, 25 + i, len_formula(f"{get_column_letter(15 + i)}{row}"))
            for i in range(3):
                ws.cell(row, 32 + i, len_formula(f"{get_column_letter(22 + i)}{row}"))
            # AU AV AW
            ws.cell(row, 47, LANDING_UTM)
            ws.cell(row, 48, DISPLAY)
            ws.cell(row, 49, "Россия")
            # BA BB — для новой кампании оставляем пусто
            ws.cell(row, 53, None)
            ws.cell(row, 54, None)
            # BC BD BE — 8 быстрых ссылок (на первой строке группы; иначе наследуются)
            if first_in_group:
                ws.cell(row, 55, sitelink_titles)
                ws.cell(row, 56, sitelink_descs)
                ws.cell(row, 57, sitelink_urls)
                ws.cell(row, 64, CLARIFICATIONS)
            else:
                ws.cell(row, 55, None)
                ws.cell(row, 56, None)
                ws.cell(row, 57, None)
                ws.cell(row, 64, None)
            # BM минус на группу
            ws.cell(row, 65, GROUP_MINUS)
            row += 1

    # --- sheet: мусорные площадки ---
    if "Мусорные площадки РСЯ" in wb.sheetnames:
        del wb["Мусорные площадки РСЯ"]
    js = wb.create_sheet("Мусорные площадки РСЯ", 1)
    js["A1"] = "Мусорные площадки РСЯ — запретить в ЕПК групповых прогулок"
    js["A1"].font = Font(bold=True, size=12)
    js["A2"] = (
        "Источник: аудиты кампаний «Морские прогулки» / выгрузки vitaminki21 "
        "(игры, мессенджеры, погода, площадки с кликами без конверсий)."
    )
    js["A4"] = "№"
    js["B4"] = "Площадка (как в Директе)"
    js["C4"] = "Комментарий"
    js["A4"].font = Font(bold=True)
    js["B4"].font = Font(bold=True)
    js["C4"].font = Font(bold=True)
    fill = PatternFill("solid", fgColor="0B3D5C")
    for col in ("A", "B", "C"):
        js[f"{col}4"].fill = fill
        js[f"{col}4"].font = Font(bold=True, color="FFFFFF")

    comments = {
        "com.imo.android.imoim": "мессенджер, 0 CR",
        "com.oakever.tiletrip": "игра, мусорные клики",
        "com.kayac.ball_run": "игра",
        "m.pogoda.yandex.ru": "погода, клики без заявок",
        "dzen.ru": "контент, низкий интент",
        "com.outfit7.mytalkingtomfriends": "Talking Tom / игры",
        "com.outfit7.talkingtomgoldrun": "Talking Tom",
        "com.outfit7.herodash": "игра",
        "com.outfit7.mytalkingangela2": "игра (были случайные конв. — всё равно режем)",
        "com.vkontakte.android": "соцсеть как РСЯ",
        "avito.ru": "классифайд как РСЯ",
        "com.avito.android": "приложение Avito",
        "sochi1.ru": "клики без конверсий",
        "otzovik.com": "отзовик",
        "nashaspravka.ru": "из аудита июля",
        "zoon.ru": "из аудита июля",
        "orgs.biz": "из аудита Алексум/Tigger",
    }
    for i, site in enumerate(sorted(set(JUNK_SITES)), 1):
        js.cell(4 + i, 1, i)
        js.cell(4 + i, 2, site)
        js.cell(4 + i, 3, comments.get(site, "мусор / игры / низкий интент"))

    js.column_dimensions["A"].width = 6
    js.column_dimensions["B"].width = 55
    js.column_dimensions["C"].width = 45

    end = 5 + len(set(JUNK_SITES))
    js.cell(end + 1, 1, "Как применить:")
    js.cell(end + 2, 1, "Директ → кампания → места показа / запрещённые площадки → вставить список из колонки B.")
    js.cell(
        end + 3,
        1,
        "yandex.ru как площадку РСЯ не включал в автозапрет (широкий охват Яндекса); "
        "смотрите отдельно после первой недели, если снова сольёт клики.",
    )

    # --- sheet: быстрые ссылки (читаемо) ---
    if "Быстрые ссылки ×8" in wb.sheetnames:
        del wb["Быстрые ссылки ×8"]
    sl = wb.create_sheet("Быстрые ссылки ×8", 2)
    sl["A1"] = "Быстрые ссылки — всегда 8 штук"
    sl["A1"].font = Font(bold=True, size=12)
    sl["A2"] = f"Посадочная: {LANDING}"
    headers = ["№", "Заголовок (≤30)", "Описание", "URL"]
    for i, h in enumerate(headers, 1):
        cell = sl.cell(4, i, h)
        cell.fill = fill
        cell.font = Font(bold=True, color="FFFFFF")
    for i in range(8):
        sl.cell(5 + i, 1, i + 1)
        sl.cell(5 + i, 2, SITELINK_TITLES[i])
        sl.cell(5 + i, 3, SITELINK_DESCS[i].replace("\n", " / "))
        sl.cell(5 + i, 4, SITELINK_URLS[i])
    sl["A14"] = "В листе «Тексты» колонки BC / BD / BE заполнены через || (формат Директа)."
    sl["A15"] = f"Телефон брони групповых: {PHONE}"
    for col, w in zip("ABCD", (6, 28, 32, 70)):
        sl.column_dimensions[col].width = w

    # note on texts sheet
    ws["A5"] = (
        f"ЕПК групповые · {LANDING} · {PHONE} · 1 800 / 2 500 ₽ · "
        "быстрые ссылки ×8 в BC–BE · мусорные площадки — лист 2"
    )

    wb.save(OUT)
    print("Wrote", OUT)
    print("rows", row - 12)
    print("junk sites", len(set(JUNK_SITES)))


if __name__ == "__main__":
    build()
